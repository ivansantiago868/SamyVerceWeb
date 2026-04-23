"""
Management command: importar_excel
Uso: python manage.py importar_excel --archivo ruta/al/archivo.xlsm
"""
import datetime
from django.core.management.base import BaseCommand
import openpyxl


def excel_serial_to_date(serial):
    """Convierte número serial de Excel a date de Python."""
    if isinstance(serial, (int, float)):
        return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(serial))).date()
    if isinstance(serial, (datetime.datetime, datetime.date)):
        return serial.date() if isinstance(serial, datetime.datetime) else serial
    return None


class Command(BaseCommand):
    help = "Importa datos desde el Excel Calculos_Contabilidad_Samy_Verce.xlsm"

    def add_arguments(self, parser):
        parser.add_argument("--archivo", type=str, required=True, help="Ruta al archivo .xlsm")

    def handle(self, *args, **options):
        # Importar modelos aquí para evitar import circular
        from apps.produccion.models import (
            Cliente, Impresora, Insumo, Gasto,
            VariablesFijas, InventarioPieza, Pedido, Venta,
        )

        ruta = options["archivo"]
        self.stdout.write(f"📂 Abriendo {ruta} ...")
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

        # ── 1. CLIENTES ──────────────────────────────────────────────
        self.stdout.write("👤 Importando Clientes...")
        ws = wb["Clientes"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_, nombre, tipo_doc, documento, email, telefono, direccion, notas = (list(row) + [None]*8)[:8]
            if id_ and nombre:
                Cliente.objects.update_or_create(
                    id=int(id_),
                    defaults={
                        "nombre":         str(nombre).strip(),
                        "tipo_documento": str(tipo_doc).strip() if tipo_doc else "",
                        "documento":      str(int(documento)) if isinstance(documento, float) else str(documento or ""),
                        "email":          str(email).strip() if email and str(email) != "N/A" else "",
                        "telefono":       str(int(telefono)) if isinstance(telefono, float) else str(telefono or ""),
                        "direccion":      str(direccion).strip() if direccion else "",
                        "notas":          str(notas).strip() if notas else "",
                    }
                )

        # ── 2. IMPRESORAS ────────────────────────────────────────────
        self.stdout.write("🖨️  Importando Impresoras...")
        ws = wb["Impresoras"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_, nombre, tipo, costo, vida_util, mant_anual = (list(row) + [None]*6)[:6]
            if id_ and nombre:
                Impresora.objects.update_or_create(
                    id=int(id_),
                    defaults={
                        "nombre":                    str(nombre).strip(),
                        "tipo":                      str(tipo).strip() if tipo else "Filamento",
                        "costo":                     costo or 0,
                        "vida_util_horas":           int(vida_util or 0),
                        "costo_mantenimiento_anual": mant_anual or 0,
                    }
                )

        # ── 3. INSUMOS ───────────────────────────────────────────────
        self.stdout.write("🧵 Importando Insumos...")
        ws = wb["INSUMOS"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_, producto, cant_ini, cant_fin, _ = (list(row) + [None]*5)[:5]
            nombre_limpio = str(producto or "").strip().lstrip("0123456789- ")
            if id_ and nombre_limpio:
                Insumo.objects.update_or_create(
                    id=int(id_),
                    defaults={
                        "producto":         nombre_limpio,
                        "cantidad_inicial": cant_ini or 0,
                        "cantidad_final":   cant_fin or 0,
                    }
                )

        # ── 4. GASTOS ────────────────────────────────────────────────
        self.stdout.write("💸 Importando Gastos...")
        ws = wb["Gastos"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            cuui, articulo, peso, costo, fecha = (list(row) + [None]*5)[:5]
            if cuui and articulo:
                Gasto.objects.update_or_create(
                    cuui=int(cuui),
                    defaults={
                        "articulo": str(articulo).strip(),
                        "peso":     peso or 0,
                        "costo":    costo or 0,
                        "fecha":    excel_serial_to_date(fecha) or datetime.date.today(),
                    }
                )

        # ── 5. VARIABLES FIJAS ───────────────────────────────────────
        self.stdout.write("⚙️  Importando Variables Fijas...")
        ws = wb["Variables Fijas"]
        datos = {}
        for row in ws.iter_rows(max_row=15, values_only=True):
            if row[0] and row[1] is not None:
                datos[str(row[0])] = row[1]
        VariablesFijas.objects.update_or_create(pk=1, defaults={
            "precio_rollo_filamento":    datos.get("Precio del rollo de filamento (COP)", 80000),
            "peso_rollo_gramos":         datos.get("Peso del rollo (gramos)", 1000),
            "costo_electricidad_kwh":    datos.get("Costo de electricidad (COP/kWh)", 850),
            "consumo_impresora_kw":      datos.get("Consumo de la impresora (kW)*", 0.15),
            "valor_hora_trabajo":        datos.get("Valor de tu hora de trabajo (COP/h)", 25000),
            "margen_ganancia":           datos.get("Margen de ganancia (%)", 0.35),
            "margen_fallos":             datos.get("Margen para fallos/impresiones fallidas (%)", 0.10),
            "costo_impresora":           datos.get("Costo de la Impresora (COP)", 1500000),
            "vida_util_horas":           int(datos.get("Vida útil estimada (Horas)", 5000)),
            "costo_mantenimiento_anual": datos.get("Costo de Mantenimiento Anual (COP", 200000),
            "horas_totales_anio":        int(datos.get("horas totales del año (8.760)", 8760)),
            "costo_empaque":             datos.get("Costo de Empaque (Caja, stickers, burbuja, etc.)", 1500),
        })

        # ── 6. INVENTARIO (piezas con costos) ────────────────────────
        self.stdout.write("📦 Importando Inventario de Piezas...")
        ws = wb["Inventario"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (list(row) + [None]*17)[:17]
            id_, nombre, peso, t_imp, t_post, c_emp, _, c_mat, c_ene, c_tpo, subtotal, fallos, depre, mant, total_real, precio_venta, _ = vals
            if id_ and nombre:
                InventarioPieza.objects.update_or_create(
                    id=int(id_),
                    defaults={
                        "nombre":                   str(nombre).strip(),
                        "peso_gramos":              peso or 0,
                        "tiempo_impresion_horas":   t_imp or 0,
                        "tiempo_postproceso_horas": t_post or 0,
                        "costo_empaque":            c_emp or 0,
                        "costo_material":           c_mat or 0,
                        "costo_energia":            c_ene or 0,
                        "costo_tiempo":             c_tpo or 0,
                        "subtotal_produccion":      subtotal or 0,
                        "amortizacion_fallos":      fallos or 0,
                        "depreciacion_pieza":       depre or 0,
                        "mantenimiento_preventivo": mant or 0,
                        "costo_total_real":         total_real or 0,
                        "precio_venta_sugerido":    precio_venta or 0,
                    }
                )

        # ── 7. PEDIDOS ───────────────────────────────────────────────
        self.stdout.write("📋 Importando Pedidos...")
        ws = wb["PEDIDOS"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (list(row) + [None]*16)[:16]
            id_, prior, cliente, producto, mat_str, cantidad, realizados, _, gr_pieza, _, precio_u, _, fecha_e, maquina, estado, desc = vals
            if id_ and cliente:
                insumo = None
                if mat_str:
                    num = str(mat_str).split("-")[0].strip()
                    insumo = Insumo.objects.filter(id=int(num)).first() if num.isdigit() else None
                impresora = Impresora.objects.filter(nombre__icontains=str(maquina).strip()).first() if maquina else None
                obj_cliente = Cliente.objects.filter(nombre__icontains=str(cliente).strip()).first() if cliente else None
                Pedido.objects.update_or_create(
                    id=int(id_),
                    defaults={
                        "prioridad":     str(prior).strip() if prior else "Medio",
                        "cliente":       obj_cliente,
                        "producto":      str(producto).strip() if producto else "",
                        "material":      insumo,
                        "cantidad":      int(cantidad or 0),
                        "realizados":    int(realizados or 0),
                        "gr_pieza":      gr_pieza or 0,
                        "precio_unidad": precio_u or 0,
                        "fecha_entrega": excel_serial_to_date(fecha_e),
                        "maquina":       impresora,
                        "estado":        str(estado).strip() if estado else "Pendiente",
                        "descripcion":   str(desc).strip() if desc else "",
                    }
                )

        # ── 8. VENTAS ────────────────────────────────────────────────
        self.stdout.write("💰 Importando Ventas...")
        ws = wb["Ventas"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            _, fecha, articulo, cantidad, pedido_id = (list(row) + [None]*5)[:5]
            if fecha and articulo:
                pedido = Pedido.objects.filter(id=int(pedido_id)).first() if pedido_id else None
                Venta.objects.get_or_create(
                    fecha=excel_serial_to_date(fecha) or datetime.date.today(),
                    articulo=str(articulo).strip(),
                    defaults={
                        "cantidad": int(cantidad or 1),
                        "pedido":   pedido,
                    }
                )

        self.stdout.write(self.style.SUCCESS("✅ Importación completada exitosamente."))
